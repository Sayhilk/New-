import requests
import fitz
from docx import Document
from openpyxl import load_workbook

def fetch_and_process_files(access_token, site_url, drive_id, folder_path):
    headers = {'Authorization': f'Bearer {access_token}'}
    endpoint = f"https://graph.microsoft.com/v1.0/sites/{site_url}/drives/{drive_id}/root:/{folder_path}:/children"
    files = requests.get(endpoint, headers=headers).json().get("value", [])

    results = {}
    for file in files:
        name = file["name"]
        download_url = file["@microsoft.graph.downloadUrl"]
        content = requests.get(download_url).content

        if name.endswith(".pdf"):
            with open(name, 'wb') as f:
                f.write(content)
            doc = fitz.open(name)
            text = "\n".join([page.get_text() for page in doc])

        elif name.endswith(".docx"):
            with open(name, 'wb') as f:
                f.write(content)
            doc = Document(name)
            text = "\n".join([p.text for p in doc.paragraphs])

        elif name.endswith(".xlsx"):
            with open(name, 'wb') as f:
                f.write(content)
            wb = load_workbook(name)
            ws = wb.active
            text = "\n".join([" | ".join([str(cell.value) for cell in row]) for row in ws.iter_rows()])

        else:
            text = "Unsupported file format."

        results[name] = text
    return results